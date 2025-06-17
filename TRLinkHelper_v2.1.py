import pandas as pd
import PySimpleGUI as sg
import os
import re
import openpyxl
import glob
import chardet
from datetime import datetime
import logging
from openpyxl.utils import get_column_letter

# 設置日誌
logging.basicConfig(filename='excel_matcher.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

def detect_encoding(file_path):
    """檢測文件編碼並打印前幾行以供調試"""
    with open(file_path, 'rb') as f:
        raw_data = f.read(1000)  # 讀取前1000字節
        result = chardet.detect(raw_data)
        encoding = result['encoding'] if result['confidence'] > 0.7 else 'utf-8'
        logging.info(f"文件 {file_path} 的前幾行（原始內容）:\n{raw_data.decode(encoding, errors='ignore').splitlines()[:5]}")
    return encoding

def read_file(path):
    """讀取CSV或Excel文件"""
    if path.endswith('.csv'):
        try:
            encoding = detect_encoding(path)
            for delimiter in [',', ';', '\t', '|']:
                try:
                    df = pd.read_csv(path, encoding=encoding, delimiter=delimiter, on_bad_lines='warn')
                    if df.shape[1] > 1:
                        logging.info(f"成功讀取CSV文件: {path}, 編碼: {encoding}, 分隔符: {delimiter}")
                        return df
                except:
                    continue
            raise Exception("無法確定CSV文件的分隔符")
        except Exception as e:
            logging.error(f"讀取CSV文件失敗: {path}, 錯誤: {str(e)}")
            raise Exception(f"無法解析CSV文件 {path}: {str(e)}")
    else:
        try:
            df = pd.read_excel(path)
            logging.info(f"成功讀取Excel文件: {path}")
            return df
        except Exception as e:
            logging.error(f"讀取Excel文件失敗: {path}, 錯誤: {str(e)}")
            raise Exception(f"無法解析Excel文件 {path}: {str(e)}")

def clean_transaction_id(trans_id):
    """清理交易ID：移除空格/特殊字符，取前14個字母數字字符"""
    s = str(trans_id).strip()
    cleaned = re.sub(r'[^A-Za-z0-9]', '', s.upper())
    return cleaned[:14] if len(cleaned) >= 14 else cleaned

def find_column_index(headers, patterns):
    """根據模式列表查找列索引"""
    for idx, header in enumerate(headers):
        header_str = str(header).strip().lower() if header else ''
        for pattern in patterns:
            if pattern in header_str:
                return idx
    return None

def process_files(tradelink_paths, cargowise_path, bill_path, output_path, window):
    """處理所有文件並生成結果"""
    try:
        window['-PROGRESS-'].update(current_count=10, max=100)
        
        # 讀取所有Tradelink文件
        all_tradelink = []
        for path in tradelink_paths:
            df = read_file(path)
            df.columns = [str(col).strip().lower() for col in df.columns]
            logging.info(f"Tradelink文件 {path} 列名: {list(df.columns)}")
            all_tradelink.append(df)
        
        if not all_tradelink:
            return False, "未找到有效的Tradelink文件"
        
        df_tradelink = pd.concat(all_tradelink, ignore_index=True)
        logging.info(f"合併後Tradelink數據行數: {len(df_tradelink)}")
        window['-PROGRESS-'].update(current_count=20, max=100)
        
        # 讀取Cargowise文件
        df_cargowise = read_file(cargowise_path)
        df_cargowise.columns = [str(col).strip().lower() for col in df_cargowise.columns]
        logging.info(f"Cargowise文件列名: {list(df_cargowise.columns)}")
        window['-PROGRESS-'].update(current_count=30, max=100)
        
        # 識別關鍵列
        tradelink_cols = {
            'udr': find_column_index(df_tradelink.columns, ['udr', '交易編號', '參考編號']),
            'house_bl': find_column_index(df_tradelink.columns, ['house bl no.', 'house bill', 'house bl']),
            'total_chrg': find_column_index(df_tradelink.columns, ['total chrg.', '總費用', '金額'])
        }
        
        cargowise_cols = {
            'trans_num': find_column_index(df_cargowise.columns, ['transaction num.', '交易編號', '參考編號']),
            'outstanding': find_column_index(df_cargowise.columns, ['outstanding amount', '未清金額', '金額']),
            'job_num': find_column_index(df_cargowise.columns, ['job number', '工作編號', 's開頭'])
        }
        
        missing_cols = [key for key, value in tradelink_cols.items() if value is None]
        if missing_cols:
            logging.error(f"Tradelink文件缺少必要列: {', '.join(missing_cols)}")
            return False, f"Tradelink文件缺少必要列: {', '.join(missing_cols)}"
        
        missing_cols = [key for key, value in cargowise_cols.items() if value is None]
        if missing_cols:
            logging.error(f"Cargowise文件缺少必要列: {', '.join(missing_cols)}")
            return False, f"Cargowise文件缺少必要列: {', '.join(missing_cols)}"
        
        df_tradelink['UDR_cleaned'] = df_tradelink.iloc[:, tradelink_cols['udr']].apply(clean_transaction_id)
        df_tradelink['House_BL'] = df_tradelink.iloc[:, tradelink_cols['house_bl']]
        df_tradelink['Total_Chrg'] = df_tradelink.iloc[:, tradelink_cols['total_chrg']]
        
        df_cargowise['Trans_Cleaned'] = df_cargowise.iloc[:, cargowise_cols['trans_num']].apply(clean_transaction_id)
        df_cargowise['Job_Num'] = df_cargowise.iloc[:, cargowise_cols['job_num']]
        df_cargowise['Outstanding'] = df_cargowise.iloc[:, cargowise_cols['outstanding']]
        
        df_cargowise_clean = df_cargowise.drop_duplicates(subset=['Trans_Cleaned'])
        window['-PROGRESS-'].update(current_count=50, max=100)
        
        merged_df = pd.merge(
            df_tradelink,
            df_cargowise_clean[['Trans_Cleaned', 'Job_Num']],
            left_on='UDR_cleaned',
            right_on='Trans_Cleaned',
            how='left'
        )
        window['-PROGRESS-'].update(current_count=70, max=100)
        
        # 處理Bill Excel文件
        wb = openpyxl.load_workbook(bill_path)
        sheet_name = "交易詳情 Transaction Details"
        
        if sheet_name not in wb.sheetnames:
            return False, f"Bill Excel中找不到工作表: {sheet_name}"
        
        sheet = wb[sheet_name]
        headers = [cell.value for cell in sheet[2]]
        logging.info(f"Bill Excel表頭: {headers}")
        
        # 擴展參考編號的模式列表
        ref_patterns = ['參考編號', 'reference number', 'ref no', 'ref. no', 'reference', '交易號', 'transaction no']
        total_patterns = ['總額', 'total amount', 'total', '金額', 'amount']
        cust_ref_patterns = ['客戶參考', 'customer ref.#5', 'customer reference', 'cust ref']
        
        ref_col_idx = find_column_index(headers, ref_patterns)
        total_col_idx = find_column_index(headers, total_patterns)
        cust_ref_col_idx = find_column_index(headers, cust_ref_patterns)
        
        if ref_col_idx is None:
            # 提示用戶選擇正確的列
            header_str = "\n".join([f"{i+1}: {h}" for i, h in enumerate(headers) if h])
            sg.popup_error(
                f"Bill Excel中找不到'參考編號'列！\n"
                f"請檢查以下表頭並更新代碼中的 ref_patterns：\n{header_str}",
                title="錯誤"
            )
            return False, f"Bill Excel中找不到'參考編號'列，請檢查表頭: {headers}"
        
        if total_col_idx is None:
            return False, f"Bill Excel中找不到'總額'列，請檢查表頭: {headers}"
        
        ref_col_idx += 1
        total_col_idx += 1
        
        if cust_ref_col_idx is None:
            last_col = sheet.max_column
            cust_ref_col_idx = last_col + 1
            sheet.cell(row=1, column=cust_ref_col_idx).value = "客戶參考 Customer Ref.#5"
        else:
            cust_ref_col_idx += 1
        
        mapping_dict = dict(zip(
            merged_df['UDR_cleaned'], 
            merged_df['Job_Num'].apply(lambda x: str(x) if not pd.isna(x) else None)
        ))
        
        update_count = 0
        total_rows = sheet.max_row - 1
        
        for row_idx in range(2, sheet.max_row + 1):
            ref_value = sheet.cell(row=row_idx, column=ref_col_idx).value
            cleaned_ref = clean_transaction_id(ref_value) if ref_value else ""
            
            if cleaned_ref in mapping_dict and mapping_dict[cleaned_ref]:
                sheet.cell(row=row_idx, column=cust_ref_col_idx).value = mapping_dict[cleaned_ref]
                update_count += 1
            
            if row_idx % 100 == 0:
                progress = 70 + int(30 * (row_idx - 1) / total_rows)
                window['-PROGRESS-'].update(current_count=progress, max=100)
        
        wb.save(output_path)
        window['-PROGRESS-'].update(current_count=100, max=100)
        
        return True, f"處理成功! 更新了 {update_count} 行數據"
    
    except Exception as e:
        logging.error(f"處理錯誤: {str(e)}", exc_info=True)
        return False, f"處理過程中出錯: {str(e)}"

# GUI設置
sg.theme('LightBlue2')
current_month = datetime.now().strftime('%B')  # Gets the full month name (e.g., June)
filename = f'Tradelinke_Statement_report_14_{current_month}.xlsx'
layout = [
    [sg.T('Tradelink文件夾 (包含CSV):'), 
     sg.Input(key='-TRADELINK_DIR-'), 
     sg.FolderBrowse()],
    
    [sg.T('Cargowise文件 (Excel/CSV):'),
     sg.Input(key='-CARGOWISE-'),
     sg.FileBrowse(file_types=(("Excel/CSV Files", "*.xlsx;*.csv"),))],
    
    [sg.T('Bill Excel文件:'),
     sg.Input(key='-BILL-'),
     sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),))],
    
    [sg.T('輸出文件:'),
     sg.Input(key='-OUTPUT-', default_text=filename),
     sg.SaveAs('保存為', key='-SAVEAS-', enable_events=True, 
               file_types=(("Excel Files", "*.xlsx"),))],
    
    [sg.ProgressBar(100, orientation='h', size=(40,20), key='-PROGRESS-', visible=False)],
    
    [sg.Button('開始處理', size=(10,1)), 
     sg.Button('退出', size=(10,1)),
     sg.T('', key='-STATUS-', size=(40,1))],
    
    [sg.Multiline('', size=(70,10), key='-LOG-', autoscroll=True, disabled=True)]
]

window = sg.Window('Excel數據匹配工具', layout, finalize=True)

while True:
    event, values = window.read()
    
    if event in (sg.WIN_CLOSED, '退出'):
        break
        
    if event == '-SAVEAS-':
        window['-OUTPUT-'].update(values['-SAVEAS-'])
        
    if event == '開始處理':
        tradelink_dir = values['-TRADELINK_DIR-']
        cargowise_path = values['-CARGOWISE-']
        bill_path = values['-BILL-']
        output_path = values['-OUTPUT-']
        
        if not all([tradelink_dir, cargowise_path, bill_path, output_path]):
            sg.popup_error('請填寫所有字段!')
            continue
            
        if not os.path.exists(tradelink_dir):
            sg.popup_error('Tradelink文件夾不存在!')
            continue
            
        if not os.path.exists(cargowise_path) or not os.path.exists(bill_path):
            sg.popup_error('輸入文件不存在!')
            continue
            
        tradelink_paths = glob.glob(os.path.join(tradelink_dir, '*.csv'))
        if not tradelink_paths:
            sg.popup_error('Tradelink文件夾中找不到CSV文件!')
            continue
            
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'
            window['-OUTPUT-'].update(output_path)
            
        if os.path.exists(output_path):
            if sg.popup_yes_no(f'輸出文件 {output_path} 已存在，是否覆蓋？') != 'Yes':
                continue
                
        window['-PROGRESS-'].update(visible=True)
        window['-STATUS-'].update('處理中...')
        window.refresh()
        
        success, message = process_files(
            tradelink_paths, 
            cargowise_path, 
            bill_path, 
            output_path, 
            window
        )
        
        window['-PROGRESS-'].update(visible=False)
        window['-STATUS-'].update(message)
        
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {'成功' if success else '錯誤'}: {message}\n"
        window['-LOG-'].update(log_entry + window['-LOG-'].get())
        
        if success:
            sg.popup('處理完成!', f'結果已保存至:\n{output_path}')
        else:
            sg.popup_error('處理失敗!', message)

window.close()