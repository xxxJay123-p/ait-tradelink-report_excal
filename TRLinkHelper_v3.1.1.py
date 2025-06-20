import pandas as pd
import os
import re
import openpyxl
import chardet
from datetime import datetime
import logging
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, NamedStyle, numbers
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import webbrowser
import shutil
import ttkbootstrap as ttk
from ttkbootstrap.constants import *

# 设置日志
logging.basicConfig(filename='trlink_helper.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

# 常量
CURRENT_MONTH = datetime.now().strftime('%B')
DEFAULT_OUTPUT = f'Tradelink_Bill_Report_{CURRENT_MONTH}.xlsx'
TDEC_DEFAULT = "16.6"

# 语言字典
LANGUAGES = {
    'en': {
        'title': 'TRLinkHelper - Trade Bill Processing Tool',
        'author': 'by xxxJay123',
        'help': 'Help Guide',
        'step1': 'Step 1: Input Files',
        'bill': 'Tradelink Bill (Excel):',
        'cargowise': 'Cargowise Data (Excel/CSV):',
        'html': 'HTML File (Optional):',
        'browse': 'Browse',
        'step2': 'Step 2: HTML Processing Settings',
        'cutoff': 'Use "After Cut-Off" Algorithm',
        'tdec': 'TDEC Transaction Amount:',
        'tdec_default': '(Default: 16.6)',
        'step3': 'Step 3: Output Settings',
        'output': 'Output File:',
        'save_as': 'Save As',
        'process': 'Start Processing',
        'logs': 'Processing Logs:',
        'error_input': 'Please provide Tradelink Bill Excel and Cargowise Report Excel!',
        'confirm_overwrite': 'Output file {} already exists. Overwrite?',
        'success': 'Completed!\nResults saved to:\n{}',
        'error': 'Error',
        'processing': 'Processing...',
        'html_processing': 'Processing HTML file... ({}/{})',
        'excel_processing': 'Processing Excel matching... ({}/{})',
        'lang_toggle': 'Switch to Chinese'
    },
    'zh': {
        'title': 'TRLinkHelper - 貿易通帳單處理工具',
        'author': 'by xxxJay123',
        'help': '使用教學',
        'step1': '步驟1: 輸入文件',
        'bill': '貿易通帳單 (Excel):',
        'cargowise': 'Cargowise 數據 (Excel/CSV):',
        'html': 'HTML 文件 (Optional):',
        'browse': '瀏覽',
        'step2': '步驟2: HTML 處理設定',
        'cutoff': '使用 "截止日期後" 算法',
        'tdec': 'TDEC 交易金額:',
        'tdec_default': '(默認: 16.6)',
        'step3': '步驟3: 輸出設定',
        'output': '輸出文件:',
        'save_as': '另存為',
        'process': '開始處理',
        'logs': '處理日誌:',
        'error_input': '請提供貿易通帳單Excel和Cargowise報表Excel!',
        'confirm_overwrite': '輸出檔案 {} 已存在，是否覆蓋?',
        'success': '完成!\n結果已儲存至:\n{}',
        'error': '錯誤',
        'processing': '處理中...',
        'html_processing': '正在處理HTML文件... ({}/{})',
        'excel_processing': '正在處理Excel匹配... ({}/{})',
        'lang_toggle': 'Switch to English'
    }
}

# ====================== 核心功能函数 ======================

def detect_encoding(file_path):
    """检测文件编码"""
    with open(file_path, 'rb') as f:
        raw_data = f.read(1000)
        result = chardet.detect(raw_data)
        return result['encoding'] if result['confidence'] > 0.7 else 'utf-8'

def clean_transaction_id(trans_id):
    """清理交易ID"""
    s = str(trans_id).strip()
    cleaned = re.sub(r'[^A-Za-z0-9]', '', s.upper())
    return cleaned[:14] if len(cleaned) >= 14 else cleaned

def find_column_index(headers, patterns):
    """根据模式列表查找列索引"""
    headers = [str(h).strip().lower() if h else "" for h in headers]
    for pattern in patterns:
        for idx, header in enumerate(headers):
            if pattern in header:
                return idx
    return None

def get_or_create_style(wb, style_name, style_config):
    """Get or create a named style in the workbook."""
    try:
        for existing_style in wb._named_styles:
            if existing_style.name == style_name:
                logging.debug(f"Found existing style: {style_name}")
                return existing_style
        
        style = NamedStyle(name=style_name)
        for attr, value in style_config.items():
            setattr(style, attr, value)
        
        try:
            wb.add_named_style(style)
            logging.debug(f"Created new style: {style_name}")
        except ValueError as e:
            if "exists already" in str(e):
                for existing_style in wb._named_styles:
                    if existing_style.name == style_name:
                        return existing_style
            else:
                raise
        
        return style
    except Exception as e:
        logging.warning(f"Error handling style {style_name}: {str(e)}")
        style = NamedStyle(name=f"{style_name}_fallback")
        for attr, value in style_config.items():
            setattr(style, attr, value)
        return style

def process_html_to_excel(html_path, bill_path, tdec_amount, is_after_cutoff, output_path, progress_callback=None):
    """Process HTML file and update Excel bill"""
    try:
        with open(html_path, 'r', encoding='utf-8') as file:
            html_data = file.read()
        
        soup = BeautifulSoup(html_data, 'html.parser')
        wb = openpyxl.load_workbook(bill_path)
        sheet_name = "交易詳情 Transaction Details"
        
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            sheet = wb[sheet_name]
            headers = [
                "Date", "Customer No.", "Reference Number", 
                "Item Description", "Ref.#1", "Ref.#2", 
                "Ref.#3", "Ref.#4", "Ref.#5", "Form Type", 
                "User", "Import/Export Declaration Charge", 
                "Clothing Levy Charge", "Penalty Charge", 
                "TDEC Transaction", "Total Amount"
            ]
            for col_idx, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_idx, value=header)
            start_row = 2
        else:
            sheet = wb[sheet_name]
            start_row = sheet.max_row + 1

        currency_columns = [12, 13, 14, 15, 16]
        currency_style = get_or_create_style(wb, 'currency_style', {
            'number_format': numbers.FORMAT_CURRENCY_USD_SIMPLE,
            'font': Font(name='NotoSansTC', size=8)
        })
        text_style = get_or_create_style(wb, 'text_style', {
            'font': Font(name='NotoSansTC', size=8)
        })
         
        transactions = []
        if is_after_cutoff:
            specific_td = None
            for td in soup.find_all('td', class_='header', colspan='6'):
                if 'EDI Transactions Since Last Charge Calculation' in td.get_text():
                    specific_td = td
                    break
                
            if not specific_td:
                logging.error("EDI Transactions table header not found")
                return False, "EDI Transactions table header not found"
                
            table = specific_td.find_parent('table')
            if not table:
                logging.error("Parent table not found")
                return False, "Parent table not found"
            
            current_row = []
            rows = table.find_all('tr')
            total_rows = len(rows)
            for idx, row in enumerate(rows):
                if row.find('td', colspan='6'):
                    ref_text = row.get_text(strip=True)
                    ref_matches = re.findall(r'Ref\. # \d+\s*-\s*([^\n<]+)', ref_text)
                    if current_row:
                        for i, ref in enumerate(ref_matches[:5]):
                            if i < 5:
                                current_row[4 + i] = ref.strip()
                else:
                    cells = [td.get_text(strip=True) for td in row.find_all(['td', 'th'])]
                    if len(cells) >= 6 and 'Amount Paidby Subscriber' not in cells and 'Transmission Date' not in cells:
                        if current_row:
                            transactions.append(current_row)
                        current_row = [''] * 16
                        current_row[0] = datetime.now().strftime('%Y/%m/%d')
                        current_row[1] = "39958258001.TL"
                        current_row[2] = cells[2] if len(cells) > 2 else ''
                        current_row[3] = "TDEC TRANSACTION"
                        current_row[10] = cells[3] if len(cells) > 3 else ''
                        current_row[11] = cells[5].replace('$', '') if len(cells) > 5 else ''
                        current_row[14] = str(tdec_amount)
                        try:
                            total = float(current_row[14]) + float(current_row[11] or 0)
                            current_row[15] = f'{total:.1f}'
                        except ValueError:
                            current_row[15] = ''
                
                # Update progress
                if progress_callback and total_rows > 0:
                    progress_callback(30 + (idx + 1) / total_rows * 30)
            
            if current_row:
                transactions.append(current_row)
            
            if not transactions:
                logging.error("No transaction data parsed")
                return False, "No transaction data parsed"
            
        else:
            table = soup.find('table', class_='soadetail')
            if not table:
                logging.error("soadetail table not found")
                return False, "soadetail table not found"
            
            temp_transactions = {}
            rows = table.find_all('tr')
            total_rows = len(rows)
            for idx, row in enumerate(rows):
                cells = [td.get_text(strip=True) for td in row.find_all('td')]
                if len(cells) != 6:
                    continue
                
                item_description = cells[0]
                date = cells[1]
                net_amount = cells[5].replace('$', '').strip()
                
                try:
                    net_amount = float(net_amount)
                except ValueError:
                    continue
                
                transaction_id_match = re.search(r'\(([A-Z0-9]A39KJEA00[A-Z0-9]+)\)', item_description)
                if not transaction_id_match:
                    continue
                transaction_id = transaction_id_match.group(1)
                
                if transaction_id not in temp_transactions:
                    temp_transactions[transaction_id] = {
                        'date': date,
                        'form_type': '',
                        'ref': '',
                        'tdec_transaction': 0.0,
                        'import_export_charge': 0.0,
                        'clothing_levy_charge': 0.0,
                        'penalty_charge': 0.0,
                        'total': 0.0
                    }
                
                if item_description.startswith('TDEC TRANSACTION'):
                    temp_transactions[transaction_id]['tdec_transaction'] = float(tdec_amount)
                elif item_description.startswith('IMPORT/EXPORT DECLARATION CHARGE'):
                    temp_transactions[transaction_id]['import_export_charge'] = net_amount
                    form_type_match = re.search(r'Form Type\s*-\s*([^<\n]+)', item_description)
                    if form_type_match:
                        temp_transactions[transaction_id]['form_type'] = form_type_match.group(1).strip()
                elif item_description.startswith('TDEC CLOTHING'):
                    temp_transactions[transaction_id]['clothing_levy_charge'] = net_amount
                elif item_description.startswith('TDEC PENALTY CHARGE'):
                    temp_transactions[transaction_id]['penalty_charge'] = net_amount
                    ref_match = re.search(r'Ref\. # 1\s*-\s*([^<\n]+)', item_description)
                    if ref_match:
                        temp_transactions[transaction_id]['ref'] = ref_match.group(1).strip()
                
                temp_transactions[transaction_id]['total'] = (
                    temp_transactions[transaction_id]['tdec_transaction'] +
                    temp_transactions[transaction_id]['import_export_charge'] +
                    temp_transactions[transaction_id]['clothing_levy_charge'] +
                    temp_transactions[transaction_id]['penalty_charge']
                )
                
                # Update progress
                if progress_callback and total_rows > 0:
                    progress_callback(30 + (idx + 1) / total_rows * 30)
            
            for transaction_id, data in temp_transactions.items():
                if data['total'] == float(tdec_amount):
                    continue
                
                if not data['date'] or len(data['date'].split('/')) != 3:
                    logging.warning(f"Invalid date format for transaction {transaction_id}: {data['date']}")
                    continue
                year, month, day = data['date'].split('/')
                
                transactions.append([
                    f"{year}/{month}/{day}",
                    "39958258001.TL",
                    transaction_id,
                    "TDEC TRANSACTION",
                    data['ref'],
                    "", "", "", "",
                    data['form_type'],
                    str(data['import_export_charge']) if data['import_export_charge'] > 0 else "",
                    str(data['clothing_levy_charge']) if data['clothing_levy_charge'] > 0 else "",
                    str(data['penalty_charge']) if data['penalty_charge'] > 0 else "",
                    "",
                    str(data['tdec_transaction']),
                    str(data['total'])
                ])
            
            if not transactions:
                logging.error("No valid transaction data parsed")
                return False, "No valid transaction data parsed"
        
        if not isinstance(transactions, list) or any(not isinstance(row, list) for row in transactions):
            logging.error(f"Invalid transactions structure: {type(transactions)}")
            return False, "Invalid transactions structure"

        total_transactions = len(transactions)
        for idx, row_data in enumerate(transactions):
            if not isinstance(row_data, list):
                logging.error(f"Unexpected row_data type: {type(row_data)}")
                return False, f"Unexpected row_data type: {type(row_data)}"
            
            if len(row_data) < 16:
                row_data.extend([''] * (16 - len(row_data)))
            
            for col_idx in range(1, 17):
                value = row_data[col_idx-1] if col_idx-1 < len(row_data) else ''
                cell = sheet.cell(row=start_row, column=col_idx, value=value)
                
                if col_idx in currency_columns:
                    cell.style = currency_style
                    try:
                        cell.value = float(value) if value else 0
                    except (ValueError, TypeError):
                        pass
                else:
                    cell.style = text_style
            start_row += 1
            
            # Update progress for writing to Excel
            if progress_callback and total_transactions > 0:
                progress_callback(60 + (idx + 1) / total_transactions * 10)
        
        wb.save(output_path)
        logging.info(f"Successfully added {len(transactions)} transaction records")
        return True, f"Successfully added {len(transactions)} transaction records"
        
    except Exception as e:
        logging.error(f"HTML processing error: {str(e)}", exc_info=True)
        return False, f"HTML processing error: {str(e)}"

def process_excel_matching(tradelink_path, cargowise_path, output_path, progress_callback=None):
    """处理Excel文件匹配"""
    try:
        if cargowise_path.endswith('.csv'):
            encoding = detect_encoding(cargowise_path)
            df_cargowise = pd.read_csv(cargowise_path, encoding=encoding, on_bad_lines='skip')
        else:
            df_cargowise = pd.read_excel(cargowise_path)
        
        cargowise_cols = {
            'trans_num': find_column_index(df_cargowise.columns, ['transaction num.', '交易編號', '參考編號', 'uxr']),
            'outstanding': find_column_index(df_cargowise.columns, ['outstanding amount', '未清金額', '金額']),
            'job_num': find_column_index(df_cargowise.columns, ['job number', '工作編號', 's開頭'])
        }
        
        missing_cols = [key for key, value in cargowise_cols.items() if value is None]
        if missing_cols:
            return False, f"Cargowise文件缺少必要列: {', '.join(missing_cols)}"
        
        df_cargowise['Trans_Cleaned'] = df_cargowise.iloc[:, cargowise_cols['trans_num']].apply(clean_transaction_id)
        df_cargowise['Job_Num'] = df_cargowise.iloc[:, cargowise_cols['job_num']].astype(str)
        df_cargowise['Outstanding'] = df_cargowise.iloc[:, cargowise_cols['outstanding']]
        
        job_mapping = df_cargowise.drop_duplicates('Trans_Cleaned').set_index('Trans_Cleaned')['Job_Num'].to_dict()
        amount_mapping = df_cargowise.drop_duplicates('Trans_Cleaned').set_index('Trans_Cleaned')['Outstanding'].to_dict()
        
        wb = openpyxl.load_workbook(tradelink_path)
        sheet_name = "交易詳情 Transaction Details"
        
        if sheet_name not in wb.sheetnames:
            return False, f"文件中找不到工作表: {sheet_name}"
        
        sheet = wb[sheet_name]
        
        headers = [cell.value for cell in sheet[2]]
        ref_col_idx = find_column_index(headers, ['參考編號', 'reference number', 'ref no', 'ref. no', 'reference', '交易號', 'transaction no'])
        total_col_idx = find_column_index(headers, ['總額', 'total amount', 'total', '金額', 'amount'])
        cust_ref_col_idx = find_column_index(headers, ['客戶參考', 'customer ref.#5', 'customer reference', 'cust ref'])
        
        if ref_col_idx is None:
            return False, "未找到'參考編號'列"
        if total_col_idx is None:
            return False, "未找到'總額'列"
        
        if cust_ref_col_idx is None:
            cust_ref_col_idx = len(headers)
            sheet.cell(row=1, column=cust_ref_col_idx+1, value="客戶參考 Customer Ref.#5")
        
        tradelink_data = []
        cargowise_data = []
        total_rows = sheet.max_row - 1  # Exclude header row
        
        for row_idx in range(2, sheet.max_row + 1):
            ref_value = sheet.cell(row=row_idx, column=ref_col_idx+1).value
            cleaned_ref = clean_transaction_id(ref_value) if ref_value else ""
            
            if cleaned_ref in job_mapping:
                sheet.cell(row=row_idx, column=cust_ref_col_idx+1, value=job_mapping[cleaned_ref])
            
            total_value = sheet.cell(row=row_idx, column=total_col_idx+1).value
            if cleaned_ref and total_value:
                tradelink_data.append({
                    "Date": sheet.cell(row=row_idx, column=1).value,
                    "Reference Number": ref_value,
                    "Total Amount": total_value
                })
            
            if cleaned_ref in amount_mapping:
                cargowise_data.append({
                    "Transaction Number": ref_value,
                    "Outstanding Amount": amount_mapping[cleaned_ref]
                })
            
            # Update progress
            if progress_callback and total_rows > 0:
                progress_callback(70 + (row_idx - 1) / total_rows * 20)
        
        if "Tradelink Matching results" in wb.sheetnames:
            del wb["Tradelink Matching results"]
        if "Cargowise Matching results" in wb.sheetnames:
            del wb["Cargowise Matching results"]

        date_style = get_or_create_style(wb, 'date_style', {
            'number_format': 'YYYY/MM/DD'
        })

        tl_sheet = wb.create_sheet("Tradelink Matching results")
        tl_sheet.append(["日期 Date", "参考编号 Reference Number", " Total Amount", "Match Status"])
        total_tradelink = len(tradelink_data)
        
        for idx, item in enumerate(tradelink_data, start=2):
            tl_sheet.append([
                item["Date"],
                item["Reference Number"],
                item["Total Amount"],
                f"=IFERROR(IF(VLOOKUP(B{idx}, 'Cargowise Matching Results'!A:B, 2, FALSE) = -C{idx}, \"Match\", \"Mismatch\"), \"Not Found\")"
            ])
            tl_sheet[f'A{idx}'].style = date_style
            # Update progress
            if progress_callback and total_tradelink > 0:
                progress_callback(90 + (idx - 1) / total_tradelink * 5)

        cw_sheet = wb.create_sheet("Cargowise Matching results")
        cw_sheet.append(["交易编号 Transaction Number", "未清金額 Outstanding Amount", "Match Status"])
        total_cargowise = len(cargowise_data)
        
        for idx, item in enumerate(cargowise_data, start=2):
            cw_sheet.append([
                item["Transaction Number"],
                item["Outstanding Amount"],
                f"=IFERROR(IF(VLOOKUP(A{idx}, 'Tradelink Matching results'!B:C, 2, FALSE) = -B{idx}, \"Match\", \"Mismatch\"), \"Not Found\")"
            ])
            # Update progress
            if progress_callback and total_cargowise > 0:
                progress_callback(95 + (idx - 1) / total_cargowise * 5)

        wb.save(output_path)
        return True, f"成功處理 {len(tradelink_data)} 筆交易記錄"
    
    except Exception as e:
        logging.error(f"Excel Matching Error: {str(e)}", exc_info=True)
        return False, f"Excel Matching Error: {str(e)}"

# ====================== GUI 界面 ======================

class TRLinkHelperApp:
    def __init__(self, root):
        self.root = root
        self.current_lang = 'zh'  # Default to Chinese
        self.root.title(LANGUAGES[self.current_lang]['title'])
        self.root.iconbitmap('trlinkhelper.ico')
        self.root.geometry("800x700")
        
        # Initialize variables
        self.bill_path = tk.StringVar()
        self.cargowise_path = tk.StringVar()
        self.html_path = tk.StringVar()
        self.output_path = tk.StringVar(value=DEFAULT_OUTPUT)
        self.tdec_amount = tk.StringVar(value=TDEC_DEFAULT)
        self.after_cutoff = tk.BooleanVar(value=True)
        
        # Store widget references for language updates
        self.widgets = {}
        self.setup_ui()
        
    def setup_ui(self):
        # Main container
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header
        self.header_frame = ttk.Frame(self.main_frame)
        self.header_frame.pack(fill=tk.X, pady=5)
        
        self.widgets['title_label'] = ttk.Label(self.header_frame, 
                                              text=LANGUAGES[self.current_lang]['title'], 
                                              font=('Arial', 16, 'bold'))
        self.widgets['title_label'].pack(side=tk.LEFT)
        
        self.widgets['author_label'] = ttk.Label(self.header_frame,
                                               text=LANGUAGES[self.current_lang]['author'],
                                               font=('Arial', 10))
        self.widgets['author_label'].pack(side=tk.LEFT, padx=10)
        
        # Language toggle and Help button
        self.widgets['lang_button'] = ttk.Button(self.header_frame, 
                                               text=LANGUAGES[self.current_lang]['lang_toggle'], 
                                               command=self.toggle_language, 
                                               style='Outline.TButton')
        self.widgets['lang_button'].pack(side=tk.RIGHT, padx=5)
        
        self.widgets['help_button'] = ttk.Button(self.header_frame, 
                                               text=LANGUAGES[self.current_lang]['help'], 
                                               command=self.open_help, 
                                               style='Outline.TButton')
        self.widgets['help_button'].pack(side=tk.RIGHT)
        
        # Separator
        ttk.Separator(self.main_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=5)
        
        # Input files section
        self.widgets['input_frame'] = ttk.LabelFrame(self.main_frame, 
                                                  text=LANGUAGES[self.current_lang]['step1'], 
                                                  padding=10)
        self.widgets['input_frame'].pack(fill=tk.X, pady=5)
        
        # Bill file
        bill_row = ttk.Frame(self.widgets['input_frame'])
        bill_row.pack(fill=tk.X, pady=5)
        self.widgets['bill_label'] = ttk.Label(bill_row, text=LANGUAGES[self.current_lang]['bill'])
        self.widgets['bill_label'].pack(side=tk.LEFT)
        ttk.Entry(bill_row, textvariable=self.bill_path, width=50).pack(side=tk.LEFT, padx=5)
        self.widgets['bill_browse'] = ttk.Button(bill_row, 
                                               text=LANGUAGES[self.current_lang]['browse'], 
                                               command=lambda: self.browse_file(self.bill_path, [("Excel Files", "*.xlsx")]))
        self.widgets['bill_browse'].pack(side=tk.LEFT)
        
        # Cargowise file
        cw_row = ttk.Frame(self.widgets['input_frame'])
        cw_row.pack(fill=tk.X, pady=5)
        self.widgets['cargowise_label'] = ttk.Label(cw_row, text=LANGUAGES[self.current_lang]['cargowise'])
        self.widgets['cargowise_label'].pack(side=tk.LEFT)
        ttk.Entry(cw_row, textvariable=self.cargowise_path, width=50).pack(side=tk.LEFT, padx=5)
        self.widgets['cargowise_browse'] = ttk.Button(cw_row, 
                                                    text=LANGUAGES[self.current_lang]['browse'], 
                                                    command=lambda: self.browse_file(self.cargowise_path, [("Excel/CSV Files", "*.xlsx;*.csv")]))
        self.widgets['cargowise_browse'].pack(side=tk.LEFT)
        
        # HTML file
        html_row = ttk.Frame(self.widgets['input_frame'])
        html_row.pack(fill=tk.X, pady=5)
        self.widgets['html_label'] = ttk.Label(html_row, text=LANGUAGES[self.current_lang]['html'])
        self.widgets['html_label'].pack(side=tk.LEFT)
        ttk.Entry(html_row, textvariable=self.html_path, width=50).pack(side=tk.LEFT, padx=5)
        self.widgets['html_browse'] = ttk.Button(html_row, 
                                               text=LANGUAGES[self.current_lang]['browse'], 
                                               command=lambda: self.browse_file(self.html_path, [("HTML Files", "*.html")]))
        self.widgets['html_browse'].pack(side=tk.LEFT)
        
        # HTML settings section
        self.widgets['html_settings_frame'] = ttk.LabelFrame(self.main_frame, 
                                                          text=LANGUAGES[self.current_lang]['step2'], 
                                                          padding=10)
        self.widgets['html_settings_frame'].pack(fill=tk.X, pady=5)
        
        self.widgets['cutoff_check'] = ttk.Checkbutton(self.widgets['html_settings_frame'], 
                                                    text=LANGUAGES[self.current_lang]['cutoff'], 
                                                    variable=self.after_cutoff)
        self.widgets['cutoff_check'].pack(anchor=tk.W, pady=2)
        
        tdec_row = ttk.Frame(self.widgets['html_settings_frame'])
        tdec_row.pack(fill=tk.X, pady=2)
        self.widgets['tdec_label'] = ttk.Label(tdec_row, text=LANGUAGES[self.current_lang]['tdec'])
        self.widgets['tdec_label'].pack(side=tk.LEFT)
        ttk.Entry(tdec_row, textvariable=self.tdec_amount, width=10).pack(side=tk.LEFT, padx=5)
        self.widgets['tdec_default_label'] = ttk.Label(tdec_row, text=LANGUAGES[self.current_lang]['tdec_default'])
        self.widgets['tdec_default_label'].pack(side=tk.LEFT)
        
        # Output settings section
        self.widgets['output_frame'] = ttk.LabelFrame(self.main_frame, 
                                                   text=LANGUAGES[self.current_lang]['step3'], 
                                                   padding=10)
        self.widgets['output_frame'].pack(fill=tk.X, pady=5)
        
        output_row = ttk.Frame(self.widgets['output_frame'])
        output_row.pack(fill=tk.X, pady=5)
        self.widgets['output_label'] = ttk.Label(output_row, text=LANGUAGES[self.current_lang]['output'])
        self.widgets['output_label'].pack(side=tk.LEFT)
        ttk.Entry(output_row, textvariable=self.output_path, width=50).pack(side=tk.LEFT, padx=5)
        self.widgets['save_as_button'] = ttk.Button(output_row, 
                                                 text=LANGUAGES[self.current_lang]['save_as'], 
                                                 command=self.save_as)
        self.widgets['save_as_button'].pack(side=tk.LEFT)
        
        # Progress bar and percentage label
        self.progress_frame = ttk.Frame(self.main_frame)
        self.progress_frame.pack(fill=tk.X, pady=10)
        self.progress_frame.pack_forget()
        
        self.progress = ttk.Progressbar(self.progress_frame, orient=tk.HORIZONTAL, length=350, mode='determinate', style='info.Horizontal.TProgressbar')
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        self.percent_label = ttk.Label(self.progress_frame, text="0%", foreground="white", background="#06525A", width=5)
        self.percent_label.pack(side=tk.LEFT)
        
        # Status label
        self.status_label = ttk.Label(self.main_frame, text="", foreground="white", anchor="center")
        self.status_label.pack(fill=tk.X, pady=5)
        
        # Action buttons
        btn_frame = ttk.Frame(self.main_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        
        self.widgets['process_button'] = ttk.Button(btn_frame, 
                                                 text=LANGUAGES[self.current_lang]['process'], 
                                                 command=self.process_files, 
                                                 style='Accent.TButton')
        self.widgets['process_button'].pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="ESC", command=self.root.quit, style='Outline.TButton').pack(side=tk.RIGHT, padx=5)
        
        # Log area
        self.widgets['log_frame'] = ttk.LabelFrame(self.main_frame, 
                                                 text=LANGUAGES[self.current_lang]['logs'], 
                                                 padding=10)
        self.widgets['log_frame'].pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.log_area = scrolledtext.ScrolledText(self.widgets['log_frame'], 
                                                wrap=tk.WORD, 
                                                width=80, 
                                                height=15,
                                                bg='#2b2b2b', 
                                                fg='white')
        self.log_area.pack(fill=tk.BOTH, expand=True)
        
        self.configure_styles()
    
    def configure_styles(self):
        style = ttk.Style()
        style.configure('Accent.TButton', foreground='#4CAF50', background='#4CAF50')
        style.configure('Outline.TButton', foreground='#ffffff')
        style.configure('info.Horizontal.TProgressbar', troughcolor='#06525A', background='#2196F3')
        style.map('Accent.TButton', 
                 background=[('active', '#45a049'), ('disabled', '#cccccc')])
        style.map('Outline.TButton',
                 background=[('active', '#3a3a3a'), ('disabled', '#cccccc')],
                 foreground=[('active', '#ffffff'), ('disabled', '#666666')])
    
    def toggle_language(self):
        self.current_lang = 'en' if self.current_lang == 'zh' else 'zh'
        self.update_ui_text()
    
    def update_ui_text(self):
        """Update the text of all widgets to the current language"""
        self.root.title(LANGUAGES[self.current_lang]['title'])
        self.widgets['title_label'].config(text=LANGUAGES[self.current_lang]['title'])
        self.widgets['author_label'].config(text=LANGUAGES[self.current_lang]['author'])
        self.widgets['lang_button'].config(text=LANGUAGES[self.current_lang]['lang_toggle'])
        self.widgets['help_button'].config(text=LANGUAGES[self.current_lang]['help'])
        self.widgets['input_frame'].config(text=LANGUAGES[self.current_lang]['step1'])
        self.widgets['bill_label'].config(text=LANGUAGES[self.current_lang]['bill'])
        self.widgets['cargowise_label'].config(text=LANGUAGES[self.current_lang]['cargowise'])
        self.widgets['html_label'].config(text=LANGUAGES[self.current_lang]['html'])
        self.widgets['bill_browse'].config(text=LANGUAGES[self.current_lang]['browse'])
        self.widgets['cargowise_browse'].config(text=LANGUAGES[self.current_lang]['browse'])
        self.widgets['html_browse'].config(text=LANGUAGES[self.current_lang]['browse'])
        self.widgets['html_settings_frame'].config(text=LANGUAGES[self.current_lang]['step2'])
        self.widgets['cutoff_check'].config(text=LANGUAGES[self.current_lang]['cutoff'])
        self.widgets['tdec_label'].config(text=LANGUAGES[self.current_lang]['tdec'])
        self.widgets['tdec_default_label'].config(text=LANGUAGES[self.current_lang]['tdec_default'])
        self.widgets['output_frame'].config(text=LANGUAGES[self.current_lang]['step3'])
        self.widgets['output_label'].config(text=LANGUAGES[self.current_lang]['output'])
        self.widgets['save_as_button'].config(text=LANGUAGES[self.current_lang]['save_as'])
        self.widgets['process_button'].config(text=LANGUAGES[self.current_lang]['process'])
        self.widgets['log_frame'].config(text=LANGUAGES[self.current_lang]['logs'])
    
    def browse_file(self, path_var, filetypes):
        filename = filedialog.askopenfilename(filetypes=filetypes)
        if filename:
            path_var.set(filename)
    
    def save_as(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=os.path.basename(self.output_path.get())
        )
        if filename:
            self.output_path.set(filename)
    
    def open_help(self):
        help_url = "https://xxxjay123-p.gitbook.io/trlinkhelper-docs/"
        webbrowser.open_new_tab(help_url)
    
    def log_message(self, message):
        self.log_area.insert(tk.END, message + "\n")
        self.log_area.see(tk.END)
        self.root.update()  # Force full UI refresh
    
    def update_progress(self, value, current=None, total=None, stage=None):
        """Update progress bar, percentage label, and status label"""
        self.progress_frame.pack(fill=tk.X, pady=10)  # Ensure frame is visible
        self.progress['value'] = value
        self.percent_label.config(text=f"{int(value)}%")
        if stage and current is not None and total is not None:
            status_text = LANGUAGES[self.current_lang][stage].format(current, total)
        else:
            status_text = LANGUAGES[self.current_lang]['processing']
        self.status_label.config(text=status_text)
        self.root.update()  # Force full UI refresh
    
    def process_files(self):
        logging.info("Starting process_files")
        if not self.bill_path.get() or not self.cargowise_path.get():
            error_msg = LANGUAGES[self.current_lang]['error_input']
            self.log_message(error_msg)
            messagebox.showerror(LANGUAGES[self.current_lang]['error'], error_msg)
            logging.error("Missing input files")
            return
            
        output_path = self.output_path.get()
        if not output_path:
            output_path = DEFAULT_OUTPUT
            self.output_path.set(output_path)
            
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'
            self.output_path.set(output_path)
            
        if os.path.exists(output_path):
            if not messagebox.askyesno(LANGUAGES[self.current_lang]['error'], 
                                     LANGUAGES[self.current_lang]['confirm_overwrite'].format(output_path)):
                logging.info("User cancelled overwrite")
                return
        
        self.progress_frame.pack(fill=tk.X, pady=10)
        self.progress['value'] = 0
        self.percent_label.config(text="0%")
        self.status_label.config(text=LANGUAGES[self.current_lang]['processing'])
        self.log_message(LANGUAGES[self.current_lang]['processing'])
        self.root.update()
        
        try:
            temp_output = "temp_bill.xlsx"
            success = True
            message = ""
            
            if self.html_path.get() and os.path.exists(self.html_path.get()):
                self.update_progress(10)
                self.log_message("Processing HTML file..." if self.current_lang == 'en' else "處理HTML文件...")
                
                def html_progress(value):
                    self.update_progress(value, stage='html_processing')
                
                success, html_msg = process_html_to_excel(
                    self.html_path.get(), 
                    self.bill_path.get(), 
                    self.tdec_amount.get(), 
                    self.after_cutoff.get(), 
                    temp_output,
                    progress_callback=html_progress
                )
                
                if not success:
                    self.status_label.config(text=f"HTML Processing Failed: {html_msg}")
                    self.log_message(f"HTML Processing Failed: {html_msg}")
                    logging.error(f"HTML processing failed: {html_msg}")
                    return
                else:
                    bill_path = temp_output
                    self.log_message(f"HTML Processing Success: {html_msg}")
            else:
                self.update_progress(60)
                shutil.copyfile(self.bill_path.get(), temp_output)
                bill_path = temp_output
                self.log_message("Skipping HTML processing step" if self.current_lang == 'en' else "跳過HTML處理步驟")
            
            self.update_progress(60)
            self.log_message("Processing Excel matching..." if self.current_lang == 'en' else "處理Excel matching...")
            
            def excel_progress(value):
                self.update_progress(value, stage='excel_processing')
            
            success, excel_msg = process_excel_matching(
                bill_path,
                self.cargowise_path.get(),
                output_path,
                progress_callback=excel_progress
            )
            
            if os.path.exists(temp_output):
                os.remove(temp_output)
            
            self.progress['value'] = 100
            self.percent_label.config(text="100%")
            self.status_label.config(text=excel_msg)
            self.log_message(f"Excel Processing: {excel_msg}")
            
            if success:
                messagebox.showinfo("Success" if self.current_lang == 'en' else "成功", 
                                  LANGUAGES[self.current_lang]['success'].format(output_path))
            else:
                messagebox.showerror(LANGUAGES[self.current_lang]['error'], 
                                   f"Processing Failed!\n{excel_msg}")
                
        except Exception as e:
            self.log_message(f"An error occurred during processing: {str(e)}")
            messagebox.showerror(LANGUAGES[self.current_lang]['error'], 
                               f"An error occurred during processing:\n{str(e)}")
            logging.error(f"Processing error: {str(e)}", exc_info=True)
        finally:
            self.progress_frame.pack_forget()

# ====================== Main Program ======================
if __name__ == '__main__':
    root = ttk.Window(themename="darkly")
    app = TRLinkHelperApp(root)
    
    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(0, weight=1)
    
    root.mainloop()
