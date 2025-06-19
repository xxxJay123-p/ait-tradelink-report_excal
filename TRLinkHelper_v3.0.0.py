import sys
import pandas as pd
import PySimpleGUI as sg
import os
import re
import openpyxl
import glob
import chardet
from datetime import datetime
import logging
import csv
from bs4 import BeautifulSoup
import json
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, NamedStyle, numbers
from openpyxl import load_workbook

# 设置日志
logging.basicConfig(filename='trlink_helper.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

# 常量
CURRENT_MONTH = datetime.now().strftime('%B')
DEFAULT_OUTPUT = f'Tradelink_Bill_Report_{CURRENT_MONTH}.xlsx'
TDEC_DEFAULT = "16.6"

# ====================== 核心功能函数 ======================

def detect_encoding(file_path):
    """检测文件编码"""
    with open(file_path, 'rb') as f:
        raw_data = f.read(1000)
        result = chardet.detect(raw_data)
        return result['encoding'] if result['confidence'] > 0.7 else 'utf-8'

def clean_transaction_id(trans_id):
    """清理交易ID"""
    s = str(trans_id).strip()
    cleaned = re.sub(r'[^A-Za-z0-9]', '', s.upper())
    return cleaned[:14] if len(cleaned) >= 14 else cleaned

def find_column_index(headers, patterns):
    """根据模式列表查找列索引"""
    headers = [str(h).strip().lower() if h else "" for h in headers]
    for pattern in patterns:
        for idx, header in enumerate(headers):
            if pattern in header:
                return idx
    return None

def process_html_to_excel(html_path, bill_path, tdec_amount, is_after_cutoff, output_path):
    """处理HTML文件并更新Excel账单"""
    try:
        with open(html_path, 'r', encoding='utf-8') as file:
            html_data = file.read()
        
        # 解析HTML
        soup = BeautifulSoup(html_data, 'html.parser')
        
        # 更新Excel文件
        wb = openpyxl.load_workbook(bill_path)
        sheet_name = "交易詳情 Transaction Details"
        
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            sheet = wb[sheet_name]
            # 添加表头
            headers = [
                "日期 Date", "客戶編號 Customer No.", "參考編號 Reference Number", 
                "項目描述 Item Description", "參考 Ref.#1", "參考 Ref.#2", 
                "參考 Ref.#3", "參考 Ref.#4", "參考 Ref.#5", "Form Type", 
                "User", "Import/Export Declaration Charge", "Clothing Levy Charge", 
                "Penalty Charge", "TDEC Transaction", "Total Amount"
            ]
            for col_idx, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_idx, value=header)
            start_row = 2
        else:
            sheet = wb[sheet_name]
            start_row = sheet.max_row + 1

        if is_after_cutoff:
            # After Cut-Off算法
            specific_td = None
            for td in soup.find_all('td', class_='header', colspan='6'):
                if 'EDI Transactions Since Last Charge Calculation' in td.get_text():
                    specific_td = td
                    break
                
            table = specific_td.find_parent('table')
            if not table:
                return False, "未找到父表格"
            
            transactions = []
            current_row = []
            
            for row in table.find_all('tr'):
                if row.find('td', colspan='6'):
                    # 处理参考信息行
                    ref_text = row.get_text(strip=True)
                    ref_matches = re.findall(r'Ref\. # \d+\s*-\s*([^\n<]+)', ref_text)
                    if current_row:
                        for i, ref in enumerate(ref_matches[:5]):
                            if i < len(current_row):
                                current_row[6 + i] = ref.strip()
                else:
                    # 处理数据行
                    cells = [td.get_text(strip=True) for td in row.find_all(['td', 'th'])]
                    if len(cells) >= 6 and 'Amount Paidby Subscriber' not in cells and 'Transmission Date' not in cells:
                        if current_row:
                            transactions.append(current_row)
                        current_row = [''] * 16  # 16列数据
                        
                        # 填充数据
                        current_row[0] = datetime.now().strftime('%Y/%m/%d')
                        current_row[1] = "39958258001.TL"
                        current_row[2] = cells[2]  # ref 
                        current_row[3] = "TDEC TRANSACTION"
                        current_row[9] = cells[3]  # User
                        current_row[10] = cells[5].replace('$', '')  # Import/Export Declaration Charge
                        current_row[14] = tdec_amount  # TDEC Transaction
                        try:
                            total = float(current_row[14]) + float(tdec_amount)
                            current_row[15] = f'{total:.1f}'
                        except ValueError:
                            current_row[15] = ''
            
            if current_row:
                transactions.append(current_row)
            
            if not transactions:
                return False, "未解析出交易数据"
            
            # 创建货币样式
            currency_style = NamedStyle(name='currency_style')
            currency_style.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            currency_style.font = Font(name='NotoSansTC', size=8)

            # 创建常规文本样式
            text_style = NamedStyle(name='text_style')
            text_style.font = Font(name='NotoSansTC', size=8)
            
            # 写入数据并应用样式
            for row_data in transactions:
                for col_idx, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=start_row, column=col_idx, value=value)
                    
                    # 应用样式
                    if col_idx in [11, 12, 13, 15, 16]:  # 这些列是货币列
                        cell.style = currency_style
                        try:
                            # 确保值是数字
                            cell.value = float(value) if value else 0
                        except (ValueError, TypeError):
                            pass
                    else:
                        cell.style = text_style
                start_row += 1
        else:
            # Before Cut-Off算法
            table = soup.find('table', class_='soadetail')
            if not table:
                return False, "未找到'soadetail'表格"
            
            transactions = {}
            for row in table.find_all('tr'):
                cells = [td.get_text(strip=True) for td in row.find_all('td')]
                if len(cells) != 6:
                    continue
                
                item_description = cells[0]
                date = cells[1]
                net_amount = cells[5].replace('$', '').strip()
                
                try:
                    net_amount = float(net_amount)
                except ValueError:
                    continue
                
                # 提取交易ID
                transaction_id_match = re.search(r'\(([A-Z0-9]A39KJEA00[A-Z0-9]+)\)', item_description)
                if not transaction_id_match:
                    continue
                transaction_id = transaction_id_match.group(1)
                
                # 初始化交易数据
                if transaction_id not in transactions:
                    transactions[transaction_id] = {
                        'date': date,
                        'form_type': '',
                        'ref': '',
                        'tdec_transaction': 0.0,
                        'import_export_charge': 0.0,
                        'clothing_levy_charge': 0.0,
                        'penalty_charge': 0.0,
                        'total': 0.0
                    }
                
                # 处理不同交易类型
                if item_description.startswith('TDEC TRANSACTION'):
                    transactions[transaction_id]['tdec_transaction'] = float(tdec_amount)
                elif item_description.startswith('IMPORT/EXPORT DECLARATION CHARGE'):
                    transactions[transaction_id]['import_export_charge'] = net_amount
                    form_type_match = re.search(r'Form Type\s*-\s*([^<\n]+)', item_description)
                    if form_type_match:
                        transactions[transaction_id]['form_type'] = form_type_match.group(1).strip()
                elif item_description.startswith('TDEC CLOTHING'):
                    transactions[transaction_id]['clothing_levy_charge'] = net_amount
                elif item_description.startswith('TDEC PENALTY CHARGE'):
                    transactions[transaction_id]['penalty_charge'] = net_amount
                    ref_match = re.search(r'Ref\. # 1\s*-\s*([^<\n]+)', item_description)
                    if ref_match:
                        transactions[transaction_id]['ref'] = ref_match.group(1).strip()
                
                # 计算总额
                transactions[transaction_id]['total'] = (
                    transactions[transaction_id]['tdec_transaction'] +
                    transactions[transaction_id]['import_export_charge'] +
                    transactions[transaction_id]['clothing_levy_charge'] +
                    transactions[transaction_id]['penalty_charge']
                )
            
            # 格式化数据
            formatted_transactions = []
            for transaction_id, data in transactions.items():
                if data['total'] == float(tdec_amount):
                    continue  # 跳过只有TDEC的交易
                
                year, month, day = data['date'].split('/')
                formatted_transactions.append([
                    f"{year}/{month}/{day}", #a 1
                    "39958258001.TL", #b 2 
                    transaction_id, #c3 
                    "TDEC TRANSACTION", #d4
                    data['ref'], #e5
                    "", #f6
                    "", #g7
                    "", #h8
                    "", #i9
                    data['form_type'], #j10
                    str(data['import_export_charge']) if data['import_export_charge'] > 0 else "", #k
                    str(data['clothing_levy_charge']) if data['clothing_levy_charge'] > 0 else "", #l
                    str(data['penalty_charge']) if data['penalty_charge'] > 0 else "", #m
                    "",
                    str(data['tdec_transaction']), #n
                    str(data['total']) #o
                ])
            
            if not formatted_transactions:
                return False, "未解析出有效交易数据"
            
            transactions = formatted_transactions
        
        # 创建货币样式
        currency_style = NamedStyle(name='currency_style')
        currency_style.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        currency_style.font = Font(name='NotoSansTC', size=8)

        # 创建常规文本样式
        text_style = NamedStyle(name='text_style')
        text_style.font = Font(name='NotoSansTC', size=8)
        
        # 写入数据并应用样式
        for row_data in transactions:
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=start_row, column=col_idx, value=value)
                
                # 应用样式
                if col_idx in [11, 12, 13, 15, 16]:  # 这些列是货币列
                    cell.style = currency_style
                    try:
                        # 确保值是数字
                        cell.value = float(value) if value else 0
                    except (ValueError, TypeError):
                        pass
                else:
                    cell.style = text_style
            start_row += 1
        
        wb.save(output_path)
        return True, f"成功添加 {len(transactions)} 条交易记录"
        
    except Exception as e:
        logging.error(f"HTML处理错误: {str(e)}", exc_info=True)
        return False, f"HTML处理失败: {str(e)}"

def process_excel_matching(tradelink_path, cargowise_path, output_path):
    """处理Excel文件匹配"""
    try:
        # 读取Cargowise文件
        if cargowise_path.endswith('.csv'):
            encoding = detect_encoding(cargowise_path)
            df_cargowise = pd.read_csv(cargowise_path, encoding=encoding, on_bad_lines='skip')
        else:
            df_cargowise = pd.read_excel(cargowise_path)
        
        # 识别关键列
        cargowise_cols = {
            'trans_num': find_column_index(df_cargowise.columns, ['transaction num.', '交易編號', '參考編號', 'uxr']),
            'outstanding': find_column_index(df_cargowise.columns, ['outstanding amount', '未清金額', '金額']),
            'job_num': find_column_index(df_cargowise.columns, ['job number', '工作編號', 's開頭'])
        }
        
        missing_cols = [key for key, value in cargowise_cols.items() if value is None]
        if missing_cols:
            return False, f"Cargowise文件缺少必要列: {', '.join(missing_cols)}"
        
        # 清理和处理Cargowise数据
        df_cargowise['Trans_Cleaned'] = df_cargowise.iloc[:, cargowise_cols['trans_num']].apply(clean_transaction_id)
        df_cargowise['Job_Num'] = df_cargowise.iloc[:, cargowise_cols['job_num']].astype(str)
        df_cargowise['Outstanding'] = df_cargowise.iloc[:, cargowise_cols['outstanding']]
        
        # 创建映射字典
        job_mapping = df_cargowise.drop_duplicates('Trans_Cleaned').set_index('Trans_Cleaned')['Job_Num'].to_dict()
        amount_mapping = df_cargowise.drop_duplicates('Trans_Cleaned').set_index('Trans_Cleaned')['Outstanding'].to_dict()
        
        # 处理Tradelink Excel文件
        wb = openpyxl.load_workbook(tradelink_path)
        sheet_name = "交易詳情 Transaction Details"
        
        if sheet_name not in wb.sheetnames:
            return False, f"文件中找不到工作表: {sheet_name}"
        
        sheet = wb[sheet_name]
        
        # 识别列
        headers = [cell.value for cell in sheet[2]]
        ref_col_idx = find_column_index(headers, ['參考編號', 'reference number', 'ref no', 'ref. no', 'reference', '交易號', 'transaction no'])
        total_col_idx = find_column_index(headers, ['總額', 'total amount', 'total', '金額', 'amount'])
        cust_ref_col_idx = find_column_index(headers, ['客戶參考', 'customer ref.#5', 'customer reference', 'cust ref'])
        
        if ref_col_idx is None:
            return False, "未找到'參考編號'列"
        if total_col_idx is None:
            return False, "未找到'總額'列"
        
        # 添加Job Number列（如果需要）
        if cust_ref_col_idx is None:
            cust_ref_col_idx = len(headers)
            sheet.cell(row=1, column=cust_ref_col_idx+1, value="客戶參考 Customer Ref.#5")
        
        # 准备匹配结果表
        tradelink_data = []
        cargowise_data = []
        
        # 处理每一行数据
        for row_idx in range(2, sheet.max_row + 1):
            ref_value = sheet.cell(row=row_idx, column=ref_col_idx+1).value
            cleaned_ref = clean_transaction_id(ref_value) if ref_value else ""
            
            # 添加Job Number
            if cleaned_ref in job_mapping:
                sheet.cell(row=row_idx, column=cust_ref_col_idx+1, value=job_mapping[cleaned_ref])
            
            # 收集匹配数据
            total_value = sheet.cell(row=row_idx, column=total_col_idx+1).value
            if cleaned_ref and total_value:
                tradelink_data.append({
                    "Date": sheet.cell(row=row_idx, column=1).value,
                    "Reference Number": ref_value,
                    "Total Amount": total_value
                })
            
            if cleaned_ref in amount_mapping:
                cargowise_data.append({
                    "Transaction Number": ref_value,
                    "Outstanding Amount": amount_mapping[cleaned_ref]
                })

        # 创建匹配结果工作表
        # First delete existing sheets if they exist
        if "Tradelink匹配结果" in wb.sheetnames:
            del wb["Tradelink匹配结果"]
        if "Cargowise匹配结果" in wb.sheetnames:
            del wb["Cargowise匹配结果"]

        # Create date style
        date_style = NamedStyle(name='date_style', number_format='YYYY/MM/DD')

        # Tradelink匹配结果表
        tl_sheet = wb.create_sheet("Tradelink匹配结果")
        tl_sheet.append(["日期 Date", "参考编号 Reference Number", " Total Amount", "Match Status"])

        for idx, item in enumerate(tradelink_data, start=2):  # Start from row 2
            tl_sheet.append([
                item["Date"],
                item["Reference Number"],
                item["Total Amount"],
                f'=IFERROR(IF(VLOOKUP(B{idx}, Cargowise匹配结果!A:B, 2, FALSE) = -C{idx}, "Match", "Mismatch"), "Not Found")'
            ])
            # Apply date style to date column
            tl_sheet[f'A{idx}'].style = date_style

        # Cargowise匹配结果表
        cw_sheet = wb.create_sheet("Cargowise匹配结果")
        cw_sheet.append(["交易编号 Transaction Number", "未清金额 Outstanding Amount", "Match Status"])

        for idx, item in enumerate(cargowise_data, start=2):  # Start from row 2
            cw_sheet.append([
                item["Transaction Number"],
                item["Outstanding Amount"],
                f'=IFERROR(IF(VLOOKUP(A{idx}, Tradelink匹配结果!B:C, 2, FALSE) = -B{idx}, "Match", "Mismatch"), "Not Found")'
            ])

        wb.save(output_path)
        return True, f"成功处理 {len(tradelink_data)} 条交易记录"
    
    except Exception as e:
        logging.error(f"Excel匹配错误: {str(e)}", exc_info=True)
        return False, f"Excel匹配失败: {str(e)}"

# ====================== GUI 界面 ======================

def create_main_window():
    """创建主界面"""
    sg.theme('LightBlue2')
    
    layout = [
        [sg.Text("TRLinkHelper - 贸易通账单处理工具", font=("Arial", 16, "bold"))],
        [sg.HorizontalSeparator()],
        
        [sg.Frame("步骤1: 输入文件", [
            [sg.Text("贸易通账单 (Excel):"), 
             sg.Input(key='-BILL-', size=(40)), 
             sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),))],
            
            [sg.Text("Cargowise 数据 (Excel/CSV):"),
             sg.Input(key='-CARGOWISE-', size=(40)),
             sg.FileBrowse(file_types=(("Excel/CSV Files", "*.xlsx;*.csv"),))],
            
            [sg.Text("HTML 文件 (可选):"),
             sg.Input(key='-HTML-', size=(40)),
             sg.FileBrowse(file_types=(("HTML Files", "*.html"),))],
        ])],
        
        [sg.Frame("步骤2: HTML 处理设置", [
            [sg.Checkbox("使用 '截止日期后' 算法", key='-AFTER_CUTOFF-', default=True)],
            [sg.Text("TDEC 交易金额:"), 
             sg.Input(key='-TDEC-', size=(10), default_text=TDEC_DEFAULT),
             sg.Text("(默认值: 16.6)")]
        ])],
        
        [sg.Frame("步骤3: 输出设置", [
            [sg.Text("输出文件:"),
             sg.Input(key='-OUTPUT-', size=(40), default_text=DEFAULT_OUTPUT),
             sg.SaveAs(file_types=(("Excel Files", "*.xlsx"),))],
        ])],
        
        [sg.ProgressBar(100, orientation='h', size=(50, 20), key='-PROGRESS-', visible=False)],
        [sg.Button('开始处理', size=(12, 1), button_color=('white', 'green')),
         sg.Button('退出', size=(12, 1), button_color=('white', 'red')),
         sg.Text('', key='-STATUS-', size=(40, 1))],
        
        [sg.Multiline('', size=(80, 10), key='-LOG-', autoscroll=True, disabled=True,
                      background_color='#f0f0f0', text_color='black')]
    ]
    
    return sg.Window('TRLinkHelper', layout, finalize=True)

# ====================== 主程序 ======================

def main():
    window = create_main_window()
    
    while True:
        event, values = window.read()
        
        if event in (sg.WIN_CLOSED, '退出'):
            break
            
        if event == '开始处理':
            # 验证输入
            bill_path = values['-BILL-']
            cargowise_path = values['-CARGOWISE-']
            html_path = values['-HTML-']
            output_path = values['-OUTPUT-']
            tdec_amount = values['-TDEC-'] or TDEC_DEFAULT
            is_after_cutoff = values['-AFTER_CUTOFF-']
            
            if not all([bill_path, cargowise_path]):
                sg.popup_error('请提供贸易通账单和Cargowise数据文件!')
                continue
                
            if not output_path:
                output_path = DEFAULT_OUTPUT
                window['-OUTPUT-'].update(output_path)
                
            if not output_path.endswith('.xlsx'):
                output_path += '.xlsx'
                window['-OUTPUT-'].update(output_path)
                
            if os.path.exists(output_path):
                if sg.popup_yes_no(f'输出文件 {output_path} 已存在，是否覆盖？') != 'Yes':
                    continue
            
            # 显示进度条
            window['-PROGRESS-'].update(visible=True, current_count=0, max=100)
            window['-STATUS-'].update('处理中...')
            window.refresh()
            
            # 处理HTML文件（如果提供）
            temp_output = "temp_bill.xlsx"
            success = True
            message = ""
            
            if html_path and os.path.exists(html_path):
                window['-PROGRESS-'].update(30)
                window.refresh()
                
                success, html_msg = process_html_to_excel(
                    html_path, 
                    bill_path, 
                    tdec_amount, 
                    is_after_cutoff, 
                    temp_output
                )
                
                if not success:
                    window['-STATUS-'].update(f"HTML处理失败: {html_msg}")
                    window['-LOG-'].update(f"HTML处理失败: {html_msg}\n", append=True)
                    continue
                else:
                    bill_path = temp_output  # 使用处理后的文件
                    window['-LOG-'].update(f"HTML处理成功: {html_msg}\n", append=True)
            else:
                # 没有HTML文件，直接复制原始文件
                import shutil
                shutil.copyfile(bill_path, temp_output)
                bill_path = temp_output
            
            # 处理Excel匹配
            window['-PROGRESS-'].update(60)
            window.refresh()
            
            success, excel_msg = process_excel_matching(
                bill_path,
                cargowise_path,
                output_path
            )
            
            # 清理临时文件
            if os.path.exists(temp_output):
                os.remove(temp_output)
            
            # 更新UI
            window['-PROGRESS-'].update(100)
            window['-STATUS-'].update(excel_msg)
            window['-LOG-'].update(f"Excel处理: {excel_msg}\n", append=True)
            
            if success:
                sg.popup('处理完成!', f'结果已保存至:\n{output_path}')
            else:
                sg.popup_error('处理失败!', excel_msg)
    
    window.close()

if __name__ == '__main__':
    main()
